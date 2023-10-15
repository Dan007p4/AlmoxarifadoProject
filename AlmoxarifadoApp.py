import pandas as pd
import streamlit as st
from streamlit_option_menu import option_menu
import streamlit_authenticator as stauth
import yaml
import xlrd
import mysql.connector
from yaml.loader import SafeLoader
from sqlalchemy import create_engine
import seaborn as sns
import matplotlib.pyplot as plt
import unidecode
import openpyxl
import datetime as dt
from datetime import datetime, timedelta
import datetime
import time
import calendar
import extra_streamlit_components as stx
from fpdf import FPDF
from streamlit_cookies_manager import EncryptedCookieManager
st.set_page_config(page_icon="🗃", page_title="Gerenciador Almoxarifado")
##FAZENDO CONEXÃO COM O DB##

connection = mysql.connector.connect(
    host="aws.connect.psdb.cloud",
    user=st.secrets["db_username"],
    passwd=st.secrets["db_password"],

    db="database",
    ssl_ca="cacert-2023-01-10.pem"
)


c = connection.cursor()


def get_manager():
    return stx.CookieManager()


cookie_manager = get_manager()
cookie = "ActualUser"


if 'Login2' not in st.session_state:
    st.session_state['Login2'] = 0

if 'new_form_menu' not in st.session_state:
    st.session_state['new_form_menu'] = 0

if 'new_form_menu_products' not in st.session_state:
    st.session_state['new_form_menu_products'] = 0


if((cookie_manager.get(cookie=cookie)) == None):
    st.session_state.Login2 = 0

if((cookie_manager.get(cookie=cookie)) == "0"):
    st.session_state.Login2 = 0

if((cookie_manager.get(cookie=cookie)) == "1"):
    st.session_state.Login2 = 1

if((cookie_manager.get(cookie=cookie)) == "2"):
    st.session_state.Login2 = 2

hide_st_style = """
             <style>
             #MainMenu {visibility: hidden;}
             footer {visibility: hidden;}
             header {visibility: hidden;}
             </style>
             """
st.markdown(hide_st_style, unsafe_allow_html=True)

if(((st.session_state.Login2 == 0) | (st.session_state.Login2 == 3))):
    cookie = "ActualUser"
    st.title("Login")

    user = st.text_input("Usuario")
    password = st.text_input("Senha", type="password")
    css = '''
            <style>
            [class="css-1li7dat effi0qh1"]{visibility: hidden;}
            </style>
            '''

    st.markdown(css, unsafe_allow_html=True)

    butt = st.button("Login")
    if butt:

        if((user == st.secrets["db_user_name"]) & (password == st.secrets["db_user_password"])):
            st.session_state.Login2 = 1
            cookie_manager.set(cookie, "1", expires_at=datetime.datetime(
                year=2024, month=2, day=2))
            time.sleep(1000)
            st.experimental_rerun()

        elif((user == st.secrets["db_adm_name"]) & (password == st.secrets["db_adm_password"])):
            st.session_state.Login2 = 2
            cookie_manager.set(cookie, "2", expires_at=datetime.datetime(
                year=2024, month=2, day=2))
            time.sleep(1000)
            st.experimental_rerun()

        else:
            st.session_state.Login2 = 3


def Clean_Names(name):
    name = str(name)
    name = unidecode.unidecode(name)
    name = name.replace(" ", '_')
    name = name.replace("/", '_')
    name = name.replace(".", '')
    return name


def NotSymbols(string):
    symbols_and_accents = [
        "a", "A", "b", "B", "c", "C", "d", "D", "e", "E",
        "f", "F", "g", "G", "h", "H", "i", "I", "j", "J",
        "k", "K", "l", "L", "m", "M", "n", "N", "o", "O",
        "p", "P", "q", "Q", "r", "R", "s", "S", "t", "T",
        "u", "U", "v", "V", "w", "W", "x", "X", "y", "Y",
        "z", "Z", "_"
    ]

    for i in string:
        if i not in symbols_and_accents:
            st.error("Simbolos não permitidos identificados no campo de texto")
            return False
    return True


def NotSymbolsDate(string):
    caracteres_permitidos = ["0123456789/"]
    for caractere in string:
        if caractere not in caracteres_permitidos[0]:
            st.error("Simbolos não permitidos identificados no campo de data")
            return False
    return True


def LogOut():
    st.session_state.Login2 = 0
    cookie_manager.set(cookie, "0", expires_at=datetime.datetime(
        year=2024, month=2, day=2))
    time.sleep(1000)

    st.experimental_rerun()


def verificar_formato_data(data_string):
    caracteres_permitidos = ["0123456789/"]
    for caractere in data_string:
        if caractere not in caracteres_permitidos[0]:
            return False
    return True


meses_dict = {
    "JANEIRO": "01",
    "FEVEREIRO": "02",
    "MARÇO": "03",
    "ABRIL": "04",
    "MAIO": "05",
    "JUNHO": "06",
    "JULHO": "07",
    "AGOSTO": "08",
    "SETEMBRO": "09",
    "OUTUBRO": "10",
    "NOVEMBRO": "11",
    "DEZEMBRO": "12"
}


if(st.session_state.Login2 == 1):
    with st.sidebar:

        selected = option_menu(
            menu_title="Menu",
            options=["Registros do almoxarifado"],
            menu_icon="border-width"
        )
    st.sidebar.image(
        "icon-Centro.jpeg", use_column_width=True)

    if selected == "Registros do almoxarifado":
        logout = st.button("Logout")
        st.divider()
        if(logout):
            LogOut()

            st.experimental_rerun()

        if(st.session_state.new_form_menu == 1):
            menu = st.button("Retornar menu")
            if menu:
                st.session_state.new_form_menu = 0
                st.experimental_rerun()

            css = '''
                    <style>
                    [class="css-1li7dat effi0qh1"]{visibility: hidden;}

                    </style>
                    '''

            st.markdown(css, unsafe_allow_html=True)
            st.title("Insira os valores para o registro no almoxarifado")
            c.execute(
                "SELECT MODELO_NOME FROM TABELA_ALMOXARIFADO_PRODUTOS;")

            list_tables = []
            tablesModelName = c.fetchall()
            for i in tablesModelName:
                value = i[0]
                list_tables.append(value)

            model = st.selectbox("Selecione o material do pedido",
                                 list_tables)
            c.execute(
                "SELECT DESCRICAO FROM TABELA_ALMOXARIFADO_PRODUTOS WHERE MODELO_NOME ='" + model+"';")

            list_tables = []
            tablesModelName = c.fetchall()
            for i in tablesModelName:
                value = i[0]
                list_tables.append(value)

            st.subheader(str(list_tables[0]))

            c.execute(
                "SELECT ESTOQUE_AUTORIZADO FROM TABELA_ALMOXARIFADO_PRODUTOS WHERE MODELO_NOME ='" + model+"';")

            list_limits = []
            tablesLimit = c.fetchall()
            for i in tablesLimit:
                value = i[0]
                list_limits.append(value)
            st.subheader("Estoque autorizado de: " + str(list_limits[0]))

            c.execute(
                "SELECT ID_ITEM FROM TABELA_ALMOXARIFADO_PRODUTOS WHERE MODELO_NOME = '" + model+"';")
            list_ID = []
            tablesID = c.fetchall()
            for i in tablesID:
                value = i[0]
                list_ID.append(value)

            c.execute(
                "SELECT SALDO_PEDIDO FROM TABELA_ALMOXARIFADO_ESTOQUE_PEDIDOS WHERE ID_PRODUTO_PEDIDO ='" + str(list_ID[0])+"';")

            list_saldo = []
            saldoPedidos = c.fetchall()
            for i in saldoPedidos:
                value = i[0]
                list_saldo.append(value)

            maior_que_o_limite = False
            if len(list_saldo) > 1:
                st.subheader("Estoque atual de: " + str(list_saldo[0]))

            else:
                st.error("Nenhum valor no estoque atual")
            val = st.number_input("Digite o valor que vai entrar",
                                  max_value=50000, min_value=1, value=1, step=1)

            if len(list_saldo) > 1:
                if (val + list_saldo[0]) > list_limits[0]:
                    st.error(
                        "Este valor somado ao estoque atual é maior que o estoque autorizado")
            else:
                maior_que_o_limite = True

            dtRecive = st.text_input(
                "Digite a data de recebimento", placeholder="Digite na seguinte formatação : 20/05/2023", max_chars=10)
            NotSymbolsDate(dtRecive)
            dtRecive = dtRecive[6:11]+dtRecive[2:7]+dtRecive[0:1]
            dtRecive = dtRecive.replace("/", "-")

            dtExpire = st.text_input(
                "Digite a data de vencimento", placeholder="Digite na seguinte formatação : 20/05/2023", max_chars=10)
            NotSymbolsDate(dtExpire)
            dtExpire = dtExpire[6:11]+dtExpire[2:7]+dtExpire[0:1]
            dtExpire = dtExpire.replace("/", "-")

            if verificar_formato_data(dtExpire) == False or verificar_formato_data(dtRecive) == False:
                st.warning("Data de recebimento ou de vencimento incorreta")

            if((dtRecive != "") & (len(dtRecive) == 10) & (len(dtExpire) == 10) & (dtExpire != "") & (val > 0) & (val != None) & (verificar_formato_data(dtExpire) != False) & (verificar_formato_data(dtRecive) != False)) & (maior_que_o_limite == True):
                send = st.button("Enviar")
                if send:
                    c.execute(
                        "SELECT ID_ITEM FROM TABELA_ALMOXARIFADO_PRODUTOS WHERE MODELO_NOME = '" + model+"';")
                    st.write("Lançamento concluído")

                    list_ID = []
                    tablesID = c.fetchall()
                    for i in tablesID:
                        value = i[0]
                        list_ID.append(value)
                    # st.text("INSERT INTO TABELA_ALMOXARIFADO_TRANSACOES( DATA_TRANSACAO, DATA_VENCIMENTO,VALOR,ID_PRODUTO) VALUES (STR_TO_DATE('" +
                    #         dtRecive+"','%Y-%m-%d'),STR_TO_DATE('"+dtExpire+"','%Y-%m-%d'),'"+str(val)+"','" + str(list_ID[0])+"');")

                    c.execute("INSERT INTO TABELA_ALMOXARIFADO_TRANSACOES( DATA_TRANSACAO, DATA_VENCIMENTO,VALOR,ID_PRODUTO) VALUES (STR_TO_DATE('" +
                              dtRecive+"','%Y-%m-%d'),STR_TO_DATE('"+dtExpire+"','%Y-%m-%d'),'"+str(val)+"','" + str(list_ID[0])+"');")

                    connection.commit()

                    # st.text("SELECT ID_TRANSACAO FROM TABELA_ALMOXARIFADO_TRANSACOES WHERE ID_PRODUTO = " +
                    #         str(list_ID[0])+"AND DATA_VENCIMENTO ='"+dtExpire+"';")
                    c.execute(
                        "SELECT ID_TRANSACAO FROM TABELA_ALMOXARIFADO_TRANSACOES WHERE ID_PRODUTO = " + str(list_ID[0])+" AND DATA_VENCIMENTO ='"+dtExpire+"';")

                    list_trans = []
                    tablesTransacao = c.fetchall()
                    for i in tablesTransacao:
                        value = i[0]
                        list_trans.append(value)

                    if len(list_trans) <= 1:

                        c.execute("INSERT INTO TABELA_ALMOXARIFADO_ESTOQUE_PEDIDOS(SALDO_PEDIDO,VENCIMENTO_PEDIDO,ID_TRANSACAO,ID_PRODUTO_PEDIDO) VALUES ('" +
                                  str(val) + "',STR_TO_DATE('"+dtExpire+"','%Y-%m-%d'),'"+str(list_trans[0])+"','" + str(list_ID[0])+"');")

                        connection.commit()

                        c.execute(
                            "SELECT ID_PEDIDO FROM TABELA_ALMOXARIFADO_ESTOQUE_PEDIDOS WHERE ID_PRODUTO_PEDIDO = " + str(list_ID[0])+" AND VENCIMENTO_PEDIDO = '"+dtExpire+"';")
                        list_val = []
                        tablesValue = c.fetchall()
                        for i in tablesValue:
                            value = i[0]
                            list_val.append(value)

                    else:
                        c.execute(
                            "SELECT SALDO_PEDIDO FROM TABELA_ALMOXARIFADO_ESTOQUE_PEDIDOS WHERE ID_PRODUTO_PEDIDO = " + str(list_ID[0])+" AND VENCIMENTO_PEDIDO = '"+dtExpire+"';")
                        list_val = []
                        tablesValue = c.fetchall()
                        for i in tablesValue:
                            value = i[0]
                            list_val.append(value)
                        addvalue = list_val[0] + val
                        # st.write(list_val[0])
                        # st.write(val)
                        # st.write(addvalue)
                        c.execute("UPDATE TABELA_ALMOXARIFADO_ESTOQUE_PEDIDOS  SET SALDO_PEDIDO = " +
                                  str(addvalue) + " WHERE  ID_PRODUTO_PEDIDO  = '" + str(list_ID[0])+"' AND VENCIMENTO_PEDIDO = '"+dtExpire+"';")
                        connection.commit()

        elif(st.session_state.new_form_menu == 2):
            menu = st.button("Retornar menu")
            if menu:
                st.session_state.new_form_menu = 0
                st.experimental_rerun()

            css = '''
                    <style>
                    [class="css-1li7dat effi0qh1"]{visibility: hidden;}

                    </style>
                    '''

            st.markdown(css, unsafe_allow_html=True)
            st.title("Selecione o item para dar baixa")

            c.execute(
                "SELECT MODELO_NOME FROM TABELA_ALMOXARIFADO_PRODUTOS;")

            list_tables = []
            tables = c.fetchall()
            for i in tables:
                value = i[0]
                list_tables.append(value)

            model = st.selectbox("Selecione o material do pedido",
                                 list_tables)
            # st.text(
            #     "SELECT ITEM_ID FROM TABELA_ALMOXARIFADO_PRODUTOS WHERE MODELO_NOME ='"+model + "';")
            c.execute(
                "SELECT ID_ITEM FROM TABELA_ALMOXARIFADO_PRODUTOS WHERE MODELO_NOME ='"+model + "';")

            list_ID = []
            tables = c.fetchall()
            for i in tables:
                value = i[0]
                list_ID.append(value)

            # st.text(
            #     "SELECT VENCIEMENTO_PEDIDO FROM TABELA_ALMOXARIFADO_ESTOQUE_PEDIDOS WHERE ID_PRODUTO_PEDIDO ='"+str(list_ID[0])+"';")
            c.execute(
                "SELECT VENCIMENTO_PEDIDO FROM TABELA_ALMOXARIFADO_ESTOQUE_PEDIDOS WHERE ID_PRODUTO_PEDIDO ='"+str(list_ID[0])+"';")

            list_Expire = []
            tables_expire = c.fetchall()
            for i in tables_expire:
                value = i[0].date()
                list_Expire.append(value)

            dtExpire = st.selectbox(
                "Selecione a data de vencimento", list_Expire)
            if (dtExpire != None):
                c.execute(
                    "SELECT SALDO_PEDIDO FROM TABELA_ALMOXARIFADO_ESTOQUE_PEDIDOS WHERE ID_PRODUTO_PEDIDO = '"+str(list_ID[0])+"' AND VENCIMENTO_PEDIDO ='" + str(dtExpire)+"';")
                list_val = []
                tablesValue = c.fetchall()
                for i in tablesValue:
                    value = i[0]
                    list_val.append(value)

                val = st.number_input("Digite o valor a ser debitado",
                                      max_value=5000, min_value=1, value=1, step=1)

                if(val > list_val[0]):
                    st.error(
                        "Valor a ser debitado é maior que todo o estoque existente do produto")

                if((val > 0) & (val != None) & (val < list_val[0])):

                    send = st.button("Enviar")
                    if send:
                        # st.text(
                        #     "SELECT VALOR FROM TABELA_ALMOXARIFADO_PRODUTOS WHERE MODELO_NOME = " + model+";")
                        c.execute(
                            "SELECT SALDO_PEDIDO FROM TABELA_ALMOXARIFADO_ESTOQUE_PEDIDOS WHERE ID_PRODUTO_PEDIDO = '"+str(list_ID[0])+"' AND VENCIMENTO_PEDIDO ='" + str(dtExpire)+"';")
                        list_val = []
                        tablesValue = c.fetchall()
                        for i in tablesValue:
                            value = i[0]
                            list_val.append(value)

                        c.execute(
                            "SELECT DATA_VENCIMENTO FROM TABELA_ALMOXARIFADO_TRANSACOES WHERE ID_PRODUTO = '" + str(list_ID[0])+"';")
                        list_Expire = []
                        tablesExpire = c.fetchall()
                        for i in tablesExpire:
                            value = i[0].date()
                            list_Expire.append(value)

                        c.execute(
                            "SELECT DATA_TRANSACAO FROM TABELA_ALMOXARIFADO_TRANSACOES WHERE ID_PRODUTO= '" + str(list_ID[0])+"';")
                        list_Recive = []
                        tablesRecive = c.fetchall()
                        for i in tablesRecive:
                            value = i[0].date()
                            list_Recive.append(value)

                        c.execute("INSERT INTO TABELA_ALMOXARIFADO_TRANSACOES( DATA_TRANSACAO, DATA_VENCIMENTO,VALOR,ID_PRODUTO) VALUES (STR_TO_DATE('" +
                                  str(list_Recive[0])+"','%Y-%m-%d'),STR_TO_DATE('"+str(dtExpire)+"','%Y-%m-%d'),'"+str(-val)+"','" + str(list_ID[0])+"');")

                        c.execute("UPDATE TABELA_ALMOXARIFADO_ESTOQUE_PEDIDOS SET SALDO_PEDIDO = " +
                                  str(list_val[0] - val) + " WHERE  ID_PRODUTO_PEDIDO = '" + str(list_ID[0])+"' AND VENCIMENTO_PEDIDO ='" + str(dtExpire)+"';")

                        connection.commit()

                        c.execute(
                            "SELECT VENCIMENTO_PEDIDO,ID_PRODUTO_PEDIDO FROM TABELA_ALMOXARIFADO_ESTOQUE_PEDIDOS WHERE SALDO_PEDIDO = '" + str(0)+"';")
                        tables0 = c.fetchall()
                        for i in tables0:
                            c.execute(
                                "DELETE FROM TABELA_ALMOXARIFADO_ESTOQUE_PEDIDOS WHERE VENCIMENTO_PEDIDO = '" + str(tables0[0])+"' AND ID_PRODUTO_PEDIDO ='"+str(tables0[1]) + "';")
                            connection.commit()
                        st.write("Baixa concluída")
            else:
                st.error("Nenhum valor em estoque encontrado para este produto")
        else:
            lanca = st.button("Lançar no almoxarifado")
            baixa = st.button("Dar baixa no almoxarifado")
            if lanca:
                st.session_state.new_form_menu = 1
                st.experimental_rerun()
            if baixa:
                st.session_state.new_form_menu = 2
                st.experimental_rerun()


elif(st.session_state.Login2 == 2):
    with st.sidebar:

        selected = option_menu(
            menu_title="Menu",
            options=["Gerenciar almoxarifado",
                     "Adicionar e remover produtos", "Registros do almoxarifado"],
            menu_icon="border-width"
        )
    st.sidebar.image(
        "icon-Centro.jpeg", use_column_width=True)

    if selected == "Registros do almoxarifado":
        logout = st.button("Logout")
        st.divider()
        if(logout):
            LogOut()

            st.experimental_rerun()

        if(st.session_state.new_form_menu == 1):
            menu = st.button("Retornar menu")
            if menu:
                st.session_state.new_form_menu = 0
                st.experimental_rerun()

            css = '''
                    <style>
                    [class="css-1li7dat effi0qh1"]{visibility: hidden;}

                    </style>
                    '''

            st.markdown(css, unsafe_allow_html=True)
            st.title("Insira os valores para o registro no almoxarifado")
            c.execute(
                "SELECT MODELO_NOME FROM TABELA_ALMOXARIFADO_PRODUTOS;")

            list_tables = []
            tablesModelName = c.fetchall()
            for i in tablesModelName:
                value = i[0]
                list_tables.append(value)

            model = st.selectbox("Selecione o material do pedido",
                                 list_tables)
            c.execute(
                "SELECT DESCRICAO FROM TABELA_ALMOXARIFADO_PRODUTOS WHERE MODELO_NOME ='" + model+"';")

            list_tables = []
            tablesModelName = c.fetchall()
            for i in tablesModelName:
                value = i[0]
                list_tables.append(value)

            st.subheader(str(list_tables[0]))

            c.execute(
                "SELECT ESTOQUE_AUTORIZADO FROM TABELA_ALMOXARIFADO_PRODUTOS WHERE MODELO_NOME ='" + model+"';")

            list_limits = []
            tablesLimit = c.fetchall()
            for i in tablesLimit:
                value = i[0]
                list_limits.append(value)
            st.subheader("Estoque autorizado de: " + str(list_limits[0]))

            c.execute(
                "SELECT ID_ITEM FROM TABELA_ALMOXARIFADO_PRODUTOS WHERE MODELO_NOME = '" + model+"';")
            list_ID = []
            tablesID = c.fetchall()
            for i in tablesID:
                value = i[0]
                list_ID.append(value)

            c.execute(
                "SELECT SALDO_PEDIDO FROM TABELA_ALMOXARIFADO_ESTOQUE_PEDIDOS WHERE ID_PRODUTO_PEDIDO ='" + str(list_ID[0])+"';")

            list_saldo = []
            saldoPedidos = c.fetchall()
            for i in saldoPedidos:
                value = i[0]
                list_saldo.append(value)

            maior_que_o_limite = False
            if len(list_saldo) > 1:
                st.subheader("Estoque atual de: " + str(list_saldo[0]))

            else:
                st.error("Nenhum valor no estoque atual")
            val = st.number_input("Digite o valor que vai entrar",
                                  max_value=50000, min_value=1, value=1, step=1)

            if len(list_saldo) > 1:
                if (val + list_saldo[0]) > list_limits[0]:
                    st.error(
                        "Este valor somado ao estoque atual é maior que o estoque autorizado")
            else:
                maior_que_o_limite = True

            dtRecive = st.text_input(
                "Digite a data de recebimento", placeholder="Digite na seguinte formatação : 20/05/2023", max_chars=10)
            NotSymbolsDate(dtRecive)
            dtRecive = dtRecive[6:11]+dtRecive[2:7]+dtRecive[0:1]
            dtRecive = dtRecive.replace("/", "-")

            dtExpire = st.text_input(
                "Digite a data de vencimento", placeholder="Digite na seguinte formatação : 20/05/2023", max_chars=10)
            NotSymbolsDate(dtExpire)
            dtExpire = dtExpire[6:11]+dtExpire[2:7]+dtExpire[0:1]
            dtExpire = dtExpire.replace("/", "-")

            if verificar_formato_data(dtExpire) == False or verificar_formato_data(dtRecive) == False:
                st.warning("Data de recebimento ou de vencimento incorreta")

            if((dtRecive != "") & (len(dtRecive) == 10) & (len(dtExpire) == 10) & (dtExpire != "") & (val > 0) & (val != None) & (verificar_formato_data(dtExpire) != False) & (verificar_formato_data(dtRecive) != False)) & (maior_que_o_limite == True):
                send = st.button("Enviar")
                if send:
                    c.execute(
                        "SELECT ID_ITEM FROM TABELA_ALMOXARIFADO_PRODUTOS WHERE MODELO_NOME = '" + model+"';")
                    st.write("Lançamento concluído")

                    list_ID = []
                    tablesID = c.fetchall()
                    for i in tablesID:
                        value = i[0]
                        list_ID.append(value)
                    # st.text("INSERT INTO TABELA_ALMOXARIFADO_TRANSACOES( DATA_TRANSACAO, DATA_VENCIMENTO,VALOR,ID_PRODUTO) VALUES (STR_TO_DATE('" +
                    #         dtRecive+"','%Y-%m-%d'),STR_TO_DATE('"+dtExpire+"','%Y-%m-%d'),'"+str(val)+"','" + str(list_ID[0])+"');")

                    c.execute("INSERT INTO TABELA_ALMOXARIFADO_TRANSACOES( DATA_TRANSACAO, DATA_VENCIMENTO,VALOR,ID_PRODUTO) VALUES (STR_TO_DATE('" +
                              dtRecive+"','%Y-%m-%d'),STR_TO_DATE('"+dtExpire+"','%Y-%m-%d'),'"+str(val)+"','" + str(list_ID[0])+"');")

                    connection.commit()

                    # st.text("SELECT ID_TRANSACAO FROM TABELA_ALMOXARIFADO_TRANSACOES WHERE ID_PRODUTO = " +
                    #         str(list_ID[0])+"AND DATA_VENCIMENTO ='"+dtExpire+"';")
                    c.execute(
                        "SELECT ID_TRANSACAO FROM TABELA_ALMOXARIFADO_TRANSACOES WHERE ID_PRODUTO = " + str(list_ID[0])+" AND DATA_VENCIMENTO ='"+dtExpire+"';")

                    list_trans = []
                    tablesTransacao = c.fetchall()
                    for i in tablesTransacao:
                        value = i[0]
                        list_trans.append(value)

                    if len(list_trans) <= 1:

                        c.execute("INSERT INTO TABELA_ALMOXARIFADO_ESTOQUE_PEDIDOS(SALDO_PEDIDO,VENCIMENTO_PEDIDO,ID_TRANSACAO,ID_PRODUTO_PEDIDO) VALUES ('" +
                                  str(val) + "',STR_TO_DATE('"+dtExpire+"','%Y-%m-%d'),'"+str(list_trans[0])+"','" + str(list_ID[0])+"');")

                        connection.commit()

                        c.execute(
                            "SELECT ID_PEDIDO FROM TABELA_ALMOXARIFADO_ESTOQUE_PEDIDOS WHERE ID_PRODUTO_PEDIDO = " + str(list_ID[0])+" AND VENCIMENTO_PEDIDO = '"+dtExpire+"';")
                        list_val = []
                        tablesValue = c.fetchall()
                        for i in tablesValue:
                            value = i[0]
                            list_val.append(value)

                    else:
                        c.execute(
                            "SELECT SALDO_PEDIDO FROM TABELA_ALMOXARIFADO_ESTOQUE_PEDIDOS WHERE ID_PRODUTO_PEDIDO = " + str(list_ID[0])+" AND VENCIMENTO_PEDIDO = '"+dtExpire+"';")
                        list_val = []
                        tablesValue = c.fetchall()
                        for i in tablesValue:
                            value = i[0]
                            list_val.append(value)
                        addvalue = list_val[0] + val
                        # st.write(list_val[0])
                        # st.write(val)
                        # st.write(addvalue)
                        c.execute("UPDATE TABELA_ALMOXARIFADO_ESTOQUE_PEDIDOS  SET SALDO_PEDIDO = " +
                                  str(addvalue) + " WHERE  ID_PRODUTO_PEDIDO  = '" + str(list_ID[0])+"' AND VENCIMENTO_PEDIDO = '"+dtExpire+"';")
                        connection.commit()

        elif(st.session_state.new_form_menu == 2):
            menu = st.button("Retornar menu")
            if menu:
                st.session_state.new_form_menu = 0
                st.experimental_rerun()

            css = '''
                    <style>
                    [class="css-1li7dat effi0qh1"]{visibility: hidden;}

                    </style>
                    '''

            st.markdown(css, unsafe_allow_html=True)
            st.title("Selecione o item para dar baixa")

            c.execute(
                "SELECT MODELO_NOME FROM TABELA_ALMOXARIFADO_PRODUTOS;")

            list_tables = []
            tables = c.fetchall()
            for i in tables:
                value = i[0]
                list_tables.append(value)

            model = st.selectbox("Selecione o material do pedido",
                                 list_tables)
            # st.text(
            #     "SELECT ITEM_ID FROM TABELA_ALMOXARIFADO_PRODUTOS WHERE MODELO_NOME ='"+model + "';")
            c.execute(
                "SELECT ID_ITEM FROM TABELA_ALMOXARIFADO_PRODUTOS WHERE MODELO_NOME ='"+model + "';")

            list_ID = []
            tables = c.fetchall()
            for i in tables:
                value = i[0]
                list_ID.append(value)

            # st.text(
            #     "SELECT VENCIEMENTO_PEDIDO FROM TABELA_ALMOXARIFADO_ESTOQUE_PEDIDOS WHERE ID_PRODUTO_PEDIDO ='"+str(list_ID[0])+"';")
            c.execute(
                "SELECT VENCIMENTO_PEDIDO FROM TABELA_ALMOXARIFADO_ESTOQUE_PEDIDOS WHERE ID_PRODUTO_PEDIDO ='"+str(list_ID[0])+"';")

            list_Expire = []
            tables_expire = c.fetchall()
            for i in tables_expire:
                value = i[0].date()
                list_Expire.append(value)

            dtExpire = st.selectbox(
                "Selecione a data de vencimento", list_Expire)
            if (dtExpire != None):
                c.execute(
                    "SELECT SALDO_PEDIDO FROM TABELA_ALMOXARIFADO_ESTOQUE_PEDIDOS WHERE ID_PRODUTO_PEDIDO = '"+str(list_ID[0])+"' AND VENCIMENTO_PEDIDO ='" + str(dtExpire)+"';")
                list_val = []
                tablesValue = c.fetchall()
                for i in tablesValue:
                    value = i[0]
                    list_val.append(value)

                val = st.number_input("Digite o valor a ser debitado",
                                      max_value=5000, min_value=1, value=1, step=1)

                if(val > list_val[0]):
                    st.error(
                        "Valor a ser debitado é maior que todo o estoque existente do produto")

                if((val > 0) & (val != None) & (val < list_val[0])):

                    send = st.button("Enviar")
                    if send:
                        # st.text(
                        #     "SELECT VALOR FROM TABELA_ALMOXARIFADO_PRODUTOS WHERE MODELO_NOME = " + model+";")
                        c.execute(
                            "SELECT SALDO_PEDIDO FROM TABELA_ALMOXARIFADO_ESTOQUE_PEDIDOS WHERE ID_PRODUTO_PEDIDO = '"+str(list_ID[0])+"' AND VENCIMENTO_PEDIDO ='" + str(dtExpire)+"';")
                        list_val = []
                        tablesValue = c.fetchall()
                        for i in tablesValue:
                            value = i[0]
                            list_val.append(value)

                        c.execute(
                            "SELECT DATA_VENCIMENTO FROM TABELA_ALMOXARIFADO_TRANSACOES WHERE ID_PRODUTO = '" + str(list_ID[0])+"';")
                        list_Expire = []
                        tablesExpire = c.fetchall()
                        for i in tablesExpire:
                            value = i[0].date()
                            list_Expire.append(value)

                        c.execute(
                            "SELECT DATA_TRANSACAO FROM TABELA_ALMOXARIFADO_TRANSACOES WHERE ID_PRODUTO= '" + str(list_ID[0])+"';")
                        list_Recive = []
                        tablesRecive = c.fetchall()
                        for i in tablesRecive:
                            value = i[0].date()
                            list_Recive.append(value)

                        c.execute("INSERT INTO TABELA_ALMOXARIFADO_TRANSACOES( DATA_TRANSACAO, DATA_VENCIMENTO,VALOR,ID_PRODUTO) VALUES (STR_TO_DATE('" +
                                  str(list_Recive[0])+"','%Y-%m-%d'),STR_TO_DATE('"+str(dtExpire)+"','%Y-%m-%d'),'"+str(-val)+"','" + str(list_ID[0])+"');")

                        c.execute("UPDATE TABELA_ALMOXARIFADO_ESTOQUE_PEDIDOS SET SALDO_PEDIDO = " +
                                  str(list_val[0] - val) + " WHERE  ID_PRODUTO_PEDIDO = '" + str(list_ID[0])+"' AND VENCIMENTO_PEDIDO ='" + str(dtExpire)+"';")

                        connection.commit()

                        c.execute(
                            "SELECT VENCIMENTO_PEDIDO,ID_PRODUTO_PEDIDO FROM TABELA_ALMOXARIFADO_ESTOQUE_PEDIDOS WHERE SALDO_PEDIDO = '" + str(0)+"';")
                        tables0 = c.fetchall()
                        for i in tables0:
                            c.execute(
                                "DELETE FROM TABELA_ALMOXARIFADO_ESTOQUE_PEDIDOS WHERE VENCIMENTO_PEDIDO = '" + str(tables0[0])+"' AND ID_PRODUTO_PEDIDO ='"+str(tables0[1]) + "';")
                            connection.commit()
                        st.write("Baixa concluída")
            else:
                st.error("Nenhum valor em estoque encontrado para este produto")

        else:
            lanca = st.button("Lançar no almoxarifado")
            baixa = st.button("Dar baixa no almoxarifado")
            if lanca:
                st.session_state.new_form_menu = 1
                st.experimental_rerun()
            if baixa:
                st.session_state.new_form_menu = 2
                st.experimental_rerun()

    if selected == "Gerenciar almoxarifado":

        logout = st.button("Logout")
        if(logout):
            LogOut()
        st.divider()
        st.title("Gerenciador Almoxarifado")
        c.execute(
            "SELECT MODELO_NOME FROM TABELA_ALMOXARIFADO_PRODUTOS;")

        list_tables = []
        tables = c.fetchall()
        for i in tables:
            value = i[0]
            list_tables.append(value)

        list_tables.append("GERAL")

        model = st.selectbox("Selecione o produto que quer ver o saldo",
                             list_tables)

        c.execute(
            "SELECT ID_ITEM FROM TABELA_ALMOXARIFADO_PRODUTOS WHERE MODELO_NOME ='"+model + "';")

        list_tables_ID = []
        tables = c.fetchall()
        for i in tables:
            value = i[0]
            list_tables_ID.append(value)

        if(model == "GERAL"):
            data = pd.read_sql(
                "SELECT * FROM TABELA_ALMOXARIFADO_ESTOQUE_PEDIDOS", con=connection)

            dataExpire = data[pd.to_datetime(
                data["VENCIMENTO_PEDIDO"]) <= dt.datetime.now()]

            dataSoon = data[(pd.to_datetime(data["VENCIMENTO_PEDIDO"]) >= dt.datetime.now()) & (
                pd.to_datetime(data["VENCIMENTO_PEDIDO"]) <= dt.datetime.now() + timedelta(days=7))]

            dataIndate = data[data["VENCIMENTO_PEDIDO"]
                              >= dt.datetime.now() + timedelta(days=7)]

            ExpireList = dataExpire["ID_PRODUTO_PEDIDO"].values
            list_Name_Expire = []
            for i in ExpireList:
                c.execute(
                    "SELECT MODELO_NOME FROM TABELA_ALMOXARIFADO_PRODUTOS WHERE ID_ITEM = '" + str(i)+"';")
                tablesID = c.fetchall()
                for i in tablesID:
                    value = i[0]
                    list_Name_Expire.append(value)

            ExpireListValue = dataExpire["SALDO_PEDIDO"].values
            ExpireListDate = dataExpire["VENCIMENTO_PEDIDO"].values

            list_Expire_date = []
            for i in ExpireListDate:
                list_Expire_date.append(str(i)[:10])

            text = "**ITENS VENCIDOS:**\n"
            for i in range(0, len(list_Name_Expire)):
                newText = "\n:red[● **"+str(ExpireListValue[i]) + \
                    " "+str(list_Name_Expire[i]) + " VENCIDOS EM " + \
                    str(list_Expire_date[0][8:10]+list_Expire_date[0][4:8] +
                        list_Expire_date[0][0:4]) + "**]\n"

                text = text + newText

            st.subheader(text)
            st.divider()

            dataSoonList = dataSoon["ID_PRODUTO_PEDIDO"].values
            list_Name_Soon = []
            for i in dataSoonList:
                c.execute(
                    "SELECT MODELO_NOME FROM TABELA_ALMOXARIFADO_PRODUTOS WHERE ID_ITEM = '" + str(i) + "';")
                tablesID = c.fetchall()
                for i in tablesID:
                    value = i[0]
                    list_Name_Soon.append(value)

            dataSoonValue = dataSoon["SALDO_PEDIDO"].values
            dataSoonDate = dataSoon["VENCIMENTO_PEDIDO"].values

            list_Soon_date = []
            for i in dataSoonDate:
                list_Soon_date.append(str(i)[:10])

            text_soon = "**ITENS PRÓXIMOS DO VENCIMENTO:**\n"
            for i in range(0, len(list_Name_Soon)):
                newText = "\n:orange[● **" + str(dataSoonValue[i]) + \
                    " " + str(list_Name_Soon[i]) + " PRÓXIMOS DO VENCIMENTO EM " + \
                    str(list_Soon_date[i][8:10]+list_Soon_date[i]
                        [4:8]+list_Soon_date[i][0:4]) + "**]\n"

                text_soon = text_soon + newText

            st.subheader(text_soon)
            st.divider()

            dataIndateList = dataIndate["ID_PRODUTO_PEDIDO"].values
            list_Name_Indate = []
            for i in dataIndateList:
                c.execute(
                    "SELECT MODELO_NOME FROM TABELA_ALMOXARIFADO_PRODUTOS WHERE ID_ITEM = '" + str(i) + "';")
                tablesID = c.fetchall()
                for i in tablesID:
                    value = i[0]
                    list_Name_Indate.append(value)

            dataIndateValue = dataIndate["SALDO_PEDIDO"].values
            dataIndateDate = dataIndate["VENCIMENTO_PEDIDO"].values

            list_Indate_date = []
            for i in dataIndateDate:
                list_Indate_date.append(str(i)[:10])

            text_indate = "**ITENS DENTRO DO PRAZO:**\n"
            for i in range(0, len(list_Name_Indate)):
                newText = "\n:green[● **" + str(dataIndateValue[i]) + \
                    " " + str(list_Name_Indate[i]) + " DENTRO DO PRAZO COM VENCIMENTO EM " + \
                    str(list_Indate_date[i][8:10]+list_Indate_date[i]
                        [4:8]+list_Indate_date[i][0:4]) + "**]\n"

                text_indate = text_indate + newText

            st.subheader(text_indate)
            st.divider()
            st.subheader("Lista de transações")
            filtro = st.radio("Filtro", ["Data", "Nenhum"])
            if filtro == "Data":
                mes = st.selectbox("Selecione o mês minimo", [
                    "JANEIRO",
                    "FEVEREIRO",
                    "MARÇO",
                    "ABRIL",
                    "MAIO",
                    "JUNHO",
                    "JULHO",
                    "AGOSTO",
                    "SETEMBRO",
                    "OUTUBRO",
                    "NOVEMBRO",
                    "DEZEMBRO"])
                anos = []
                anos.append(dt.date.today().year)
                if dt.date.today().year > 2023:
                    num_anos = dt.date.today().year - 2023
                    for i in range(num_anos):
                        days = float(365*(i+1))
                        previous_year = dt.date.today() - timedelta(days=float(days))
                        anos.append(previous_year.year)

                ano = st.selectbox("Selecione o ano minimo", anos)
                dias = []
                for i in range(calendar.monthrange(int(ano), int(meses_dict[mes]))[1]):
                    dias.append(i+1)
                dia = st.selectbox("Selecione o dia minimo", dias)

                mes2 = st.selectbox("Selecione o mês maximo", [
                    "JANEIRO",
                    "FEVEREIRO",
                    "MARÇO",
                    "ABRIL",
                    "MAIO",
                    "JUNHO",
                    "JULHO",
                    "AGOSTO",
                    "SETEMBRO",
                    "OUTUBRO",
                    "NOVEMBRO",
                    "DEZEMBRO"])
                ano2 = st.selectbox("Selecione o ano maximo", anos)

                dias2 = []
                for i in range(calendar.monthrange(int(ano2), int(meses_dict[mes2]))[1]):
                    dias2.append(i+1)
                dia2 = st.selectbox("Selecione o dia maximo", dias2)

                data = pd.read_sql(
                    "SELECT * FROM TABELA_ALMOXARIFADO_TRANSACOES WHERE DATA_TRANSACAO <= '"+str(ano2)+"-"+str(meses_dict[mes2])+"-"+str(dia2)+"' AND  DATA_TRANSACAO >= '"+str(ano)+"-"+str(meses_dict[mes])+"-"+str(dia)+"' ;", con=connection)

            else:
                data = pd.read_sql(
                    "SELECT * FROM TABELA_ALMOXARIFADO_TRANSACOES;", con=connection)

            if len(data["ID_PRODUTO"].values) == 0:
                st.error("Nenhum registro encontrado")

            else:
                data = pd.read_sql(
                    "SELECT * FROM TABELA_ALMOXARIFADO_TRANSACOES", con=connection)

                c.execute(
                    "SELECT * FROM TABELA_ALMOXARIFADO_PRODUTOS ;")
                AllProducts = c.fetchall()
                AllProducts_Dictionaty = {}
                for i in range(len(AllProducts)):
                    AllProducts_Dictionaty[AllProducts[i][0]] = str(
                        AllProducts[i][1])
                # data["ID_PRODUTO"] = data["ID_PRODUTO"].to_string()
                data["ID_PRODUTO"] = data["ID_PRODUTO"].replace(
                    AllProducts_Dictionaty)
                data["DATA_TRANSACAO"] = data["DATA_TRANSACAO"].dt.strftime(
                    '%d-%m-%Y %H:%M:%S')
                data["DATA_VENCIMENTO"] = data["DATA_VENCIMENTO"].dt.strftime(
                    '%d-%m-%Y %H:%M:%S')
                st.dataframe(data)
                list_unique = data['ID_PRODUTO'].unique()
                st.subheader("Total das transações do periodo por item")

                for i in list_unique:
                    st.subheader(i+":\n"+"\n:red[● **Sairam um total de "+str(abs(data[(data['ID_PRODUTO'] == i) & (data['VALOR'] < 0)]["VALOR"].sum()))+"**]\n"+"\n:green[● **Entraram um total de "+str(data[(data['ID_PRODUTO'] == i) & (data['VALOR'] > 0)]["VALOR"].sum())+"** ]"+"\n"+"\n **Saldo total de " +
                                 str(data[data['ID_PRODUTO'] == i]["VALOR"].sum())+"**")

                st.divider()

                pdf = FPDF()
                pdf.add_page()
                pdf.set_font("Arial", size=12)

                text = "ITENS VENCIDOS:\n"
                pdf.cell(200, 10, txt=text, ln=1, align="C")
                pdf.ln(2)
                for i in range(0, len(list_Name_Expire)):
                    newText = "\n-- "+str(ExpireListValue[i]) + \
                        " "+str(list_Name_Expire[i]) + " VENCIDOS EM " + \
                        str(list_Expire_date[0][8:10]+list_Expire_date[0][4:8] +
                            list_Expire_date[0][0:4]) + "\n"

                    pdf.cell(200, 10, txt=newText, ln=1, align="C")
                    pdf.ln(2)

                text_soon = "ITENS PRÓXIMOS DO VENCIMENTO:\n"
                pdf.cell(200, 10, txt=text_soon, ln=1, align="C")
                pdf.ln(2)
                for i in range(0, len(list_Name_Soon)):
                    newText = "\n-- " + str(dataSoonValue[i]) + \
                        " " + str(list_Name_Soon[i]) + " PRÓXIMOS DO VENCIMENTO EM " + \
                        str(list_Soon_date[i][8:10]+list_Soon_date[i]
                            [4:8]+list_Soon_date[i][0:4]) + "\n"
                    pdf.cell(200, 10, txt=newText, ln=1, align="C")
                    pdf.ln(2)

                text_indate = "ITENS DENTRO DO PRAZO:\n"
                pdf.cell(200, 10, txt=text_indate, ln=1, align="C")
                pdf.ln(2)
                for i in range(0, len(list_Name_Indate)):
                    newText = "\n-- " + str(dataIndateValue[i]) + \
                        " " + str(list_Name_Indate[i]) + " DENTRO DO PRAZO COM VENCIMENTO EM " + \
                        str(list_Indate_date[i][8:10]+list_Indate_date[i]
                            [4:8]+list_Indate_date[i][0:4]) + "\n"

                    pdf.cell(200, 10, txt=newText, ln=1, align="C")
                    pdf.ln(2)

                pdf.output("example.pdf")

                with open("example.pdf", "rb") as f:
                    st.download_button(
                        label="Fazer dowload do relatorio em PDF",
                        data=f,
                        file_name='relatorio.pdf'
                    )

        else:

            c.execute(
                "SELECT ID_ITEM FROM TABELA_ALMOXARIFADO_PRODUTOS WHERE MODELO_NOME = '" + model+"';")
            list_ID = []
            tablesID = c.fetchall()
            for i in tablesID:
                value = i[0]
                list_ID.append(value)

            data = pd.read_sql(
                "SELECT * FROM TABELA_ALMOXARIFADO_ESTOQUE_PEDIDOS WHERE ID_PRODUTO_PEDIDO = "+str(list_ID[0])+";", con=connection)
            dataExpire = data[pd.to_datetime(
                data["VENCIMENTO_PEDIDO"]) <= dt.datetime.now()]

            dataSoon = data[(pd.to_datetime(data["VENCIMENTO_PEDIDO"]) >= dt.datetime.now()) & (
                pd.to_datetime(data["VENCIMENTO_PEDIDO"]) <= dt.datetime.now() + timedelta(days=7))]

            dataIndate = data[data["VENCIMENTO_PEDIDO"]
                              >= dt.datetime.now() + timedelta(days=7)]

            ExpireList = dataExpire["ID_PRODUTO_PEDIDO"].values
            list_Name_Expire = []
            for i in ExpireList:
                c.execute(
                    "SELECT MODELO_NOME FROM TABELA_ALMOXARIFADO_PRODUTOS WHERE ID_ITEM = '" + str(i)+"';")
                tablesID = c.fetchall()
                for i in tablesID:
                    value = i[0]
                    list_Name_Expire.append(value)

            ExpireListValue = dataExpire["SALDO_PEDIDO"].values
            ExpireListDate = dataExpire["VENCIMENTO_PEDIDO"].values

            list_Expire_date = []
            for i in ExpireListDate:
                list_Expire_date.append(str(i)[:10])

            text = "**ITENS VENCIDOS:**\n"
            for i in range(0, len(list_Name_Expire)):
                newText = "\n:red[● **"+str(ExpireListValue[i]) + \
                    " "+str(list_Name_Expire[i]) + " VENCIDOS EM " + \
                    str(list_Expire_date[0][8:10]+list_Expire_date[0][4:8] +
                        list_Expire_date[0][0:4]) + "**]\n"

                text = text + newText

            st.subheader(text)
            st.divider()

            dataSoonList = dataSoon["ID_PRODUTO_PEDIDO"].values
            list_Name_Soon = []
            for i in dataSoonList:
                c.execute(
                    "SELECT MODELO_NOME FROM TABELA_ALMOXARIFADO_PRODUTOS WHERE ID_ITEM = '" + str(i) + "';")
                tablesID = c.fetchall()
                for i in tablesID:
                    value = i[0]
                    list_Name_Soon.append(value)

            dataSoonValue = dataSoon["SALDO_PEDIDO"].values
            dataSoonDate = dataSoon["VENCIMENTO_PEDIDO"].values

            list_Soon_date = []
            for i in dataSoonDate:
                list_Soon_date.append(str(i)[:10])

            text_soon = "**ITENS PRÓXIMOS DO VENCIMENTO:**\n"
            for i in range(0, len(list_Name_Soon)):
                newText = "\n:orange[● **" + str(dataSoonValue[i]) + \
                    " " + str(list_Name_Soon[i]) + " PRÓXIMOS DO VENCIMENTO EM " + \
                    str(list_Soon_date[i][8:10]+list_Soon_date[i]
                        [4:8]+list_Soon_date[i][0:4]) + "**]\n"

                text_soon = text_soon + newText

            st.subheader(text_soon)
            st.divider()

            dataIndateList = dataIndate["ID_PRODUTO_PEDIDO"].values
            list_Name_Indate = []
            for i in dataIndateList:
                c.execute(
                    "SELECT MODELO_NOME FROM TABELA_ALMOXARIFADO_PRODUTOS WHERE ID_ITEM = '" + str(i) + "';")
                tablesID = c.fetchall()
                for i in tablesID:
                    value = i[0]
                    list_Name_Indate.append(value)

            dataIndateValue = dataIndate["SALDO_PEDIDO"].values
            dataIndateDate = dataIndate["VENCIMENTO_PEDIDO"].values

            list_Indate_date = []
            for i in dataIndateDate:
                list_Indate_date.append(str(i)[:10])

            text_indate = "**ITENS DENTRO DO PRAZO:**\n"
            for i in range(0, len(list_Name_Indate)):
                newText = "\n:green[● **" + str(dataIndateValue[i]) + \
                    " " + str(list_Name_Indate[i]) + " DENTRO DO PRAZO EM " + \
                    str(list_Indate_date[i][8:10]+list_Indate_date[i]
                        [4:8]+list_Indate_date[i][0:4]) + "**]\n"

                text_indate = text_indate + newText

            st.subheader(text_indate)
            st.divider()
            st.subheader("Lista de transações")
            filtro = st.radio("Filtro", ["Data", "Nenhum"])
            if filtro == "Data":
                mes = st.selectbox("Selecione o mês minimo", [
                    "JANEIRO",
                    "FEVEREIRO",
                    "MARÇO",
                    "ABRIL",
                    "MAIO",
                    "JUNHO",
                    "JULHO",
                    "AGOSTO",
                    "SETEMBRO",
                    "OUTUBRO",
                    "NOVEMBRO",
                    "DEZEMBRO"])
                anos = []
                anos.append(dt.date.today().year)
                if dt.date.today().year > 2023:
                    num_anos = dt.date.today().year - 2023
                    for i in range(num_anos):
                        days = float(365*(i+1))
                        previous_year = dt.date.today() - timedelta(days=float(days))
                        anos.append(previous_year.year)

                ano = st.selectbox("Selecione o ano minimo", anos)
                dias = []
                for i in range(calendar.monthrange(int(ano), int(meses_dict[mes]))[1]):
                    dias.append(i+1)
                dia = st.selectbox("Selecione o dia minimo", dias)

                mes2 = st.selectbox("Selecione o mês maximo", [
                    "JANEIRO",
                    "FEVEREIRO",
                    "MARÇO",
                    "ABRIL",
                    "MAIO",
                    "JUNHO",
                    "JULHO",
                    "AGOSTO",
                    "SETEMBRO",
                    "OUTUBRO",
                    "NOVEMBRO",
                    "DEZEMBRO"])
                ano2 = st.selectbox("Selecione o ano maximo", anos)

                dias2 = []
                for i in range(calendar.monthrange(int(ano2), int(meses_dict[mes2]))[1]):
                    dias2.append(i+1)
                dia2 = st.selectbox("Selecione o dia maximo", dias2)

                data = pd.read_sql(
                    "SELECT * FROM TABELA_ALMOXARIFADO_TRANSACOES WHERE ID_PRODUTO = '" + str(list_tables_ID[0])+"' AND  DATA_TRANSACAO <= '"+str(ano2)+"-"+str(meses_dict[mes2])+"-"+str(dia2)+"' AND  DATA_TRANSACAO >= '"+str(ano)+"-"+str(meses_dict[mes])+"-"+str(dia)+"' ;", con=connection)

            else:
                data = pd.read_sql(
                    "SELECT * FROM TABELA_ALMOXARIFADO_TRANSACOES WHERE ID_PRODUTO = '" + str(list_tables_ID[0])+"';", con=connection)

            if len(data["ID_PRODUTO"].values) == 0:
                st.error("Nenhum registro encontrado")

            else:
                c.execute(
                    "SELECT * FROM TABELA_ALMOXARIFADO_PRODUTOS ;")
                AllProducts = c.fetchall()
                AllProducts_Dictionaty = {}
                for i in range(len(AllProducts)):
                    AllProducts_Dictionaty[AllProducts[i][0]] = str(
                        AllProducts[i][1])
                # data["ID_PRODUTO"] = data["ID_PRODUTO"].to_string()
                data["ID_PRODUTO"] = data["ID_PRODUTO"].replace(
                    AllProducts_Dictionaty)
                data["DATA_TRANSACAO"] = data["DATA_TRANSACAO"].dt.strftime(
                    '%d-%m-%Y %H:%M:%S')
                data["DATA_VENCIMENTO"] = data["DATA_VENCIMENTO"].dt.strftime(
                    '%d-%m-%Y %H:%M:%S')

                st.dataframe(data)
                list_unique = data['ID_PRODUTO'].unique()
                st.subheader("Total das transações do periodo por item")

                for i in list_unique:
                    st.subheader(i+":\n"+"\n:red[● **Sairam um total de "+str(abs(data[(data['ID_PRODUTO'] == i) & (data['VALOR'] < 0)]["VALOR"].sum()))+"**]\n"+"\n:green[● **Entraram um total de "+str(data[(data['ID_PRODUTO'] == i) & (data['VALOR'] > 0)]["VALOR"].sum())+"** ]"+"\n"+"\n **Saldo total de " +
                                 str(data[data['ID_PRODUTO'] == i]["VALOR"].sum())+"**")

                st.divider()

                pdf = FPDF()
                pdf.add_page()
                pdf.set_font("Arial", size=12)

                text = "ITENS VENCIDOS:\n"
                pdf.cell(200, 10, txt=text, ln=1, align="C")
                pdf.ln(2)
                for i in range(0, len(list_Name_Expire)):
                    newText = "\n-- "+str(ExpireListValue[i]) + \
                        " "+str(list_Name_Expire[i]) + " VENCIDOS EM " + \
                        str(list_Expire_date[0][8:10]+list_Expire_date[0][4:8] +
                            list_Expire_date[0][0:4]) + "\n"

                    pdf.cell(200, 10, txt=newText, ln=1, align="C")
                    pdf.ln(2)

                text_soon = "ITENS PRÓXIMOS DO VENCIMENTO:\n"
                pdf.cell(200, 10, txt=text_soon, ln=1, align="C")
                pdf.ln(2)
                for i in range(0, len(list_Name_Soon)):
                    newText = "\n-- " + str(dataSoonValue[i]) + \
                        " " + str(list_Name_Soon[i]) + " PRÓXIMOS DO VENCIMENTO EM " + \
                        str(list_Soon_date[i][8:10]+list_Soon_date[i]
                            [4:8]+list_Soon_date[i][0:4]) + "\n"
                    pdf.cell(200, 10, txt=newText, ln=1, align="C")
                    pdf.ln(2)

                text_indate = "ITENS DENTRO DO PRAZO:\n"
                pdf.cell(200, 10, txt=text_indate, ln=1, align="C")
                pdf.ln(2)
                for i in range(0, len(list_Name_Indate)):
                    newText = "\n-- " + str(dataIndateValue[i]) + \
                        " " + str(list_Name_Indate[i]) + " DENTRO DO PRAZO COM VENCIMENTO EM " + \
                        str(list_Indate_date[i][8:10]+list_Indate_date[i]
                            [4:8]+list_Indate_date[i][0:4]) + "\n"

                    pdf.cell(200, 10, txt=newText, ln=1, align="C")
                    pdf.ln(2)

                pdf.output("example.pdf")

                with open("example.pdf", "rb") as f:
                    st.download_button(
                        label="Fazer dowload do relatorio em PDF",
                        data=f,
                        file_name='relatorio.pdf'
                    )

    if selected == "Adicionar e remover produtos":
        logout = st.button("Logout")
        if(logout):
            LogOut()
        st.divider()

        if st.session_state.new_form_menu_products == 4:
            menu = st.button("Retornar menu")
            if menu:
                st.session_state.new_form_menu_products = 0
                st.experimental_rerun()
            st.title("Adicionar descrição")
            desc = st.text_input("Insira o nome da descrição")
            st.button("Salvar descrição")

        if st.session_state.new_form_menu_products == 2:
            menu = st.button("Retornar menu")
            if menu:
                st.session_state.new_form_menu_products = 0
                st.experimental_rerun()

            st.title("Remover produto")

            c.execute(
                "SELECT MODELO_NOME FROM TABELA_ALMOXARIFADO_PRODUTOS;")
            list_tables = []
            tablesModelName = c.fetchall()
            for i in tablesModelName:
                value = i[0]
                list_tables.append(value)

            model = st.selectbox("Selecione o material do pedido",
                                 list_tables)
            delete = st.button("Deletar")
            if delete:
                c.execute(
                    "DELETE FROM TABELA_ALMOXARIFADO_PRODUTOS WHERE MODELO_NOME = '" + model+"';")
                connection.commit()
                st.write(":green[Removido com sucesso!]")

        if st.session_state.new_form_menu_products == 3:
            menu = st.button("Retornar menu")
            if menu:
                st.session_state.new_form_menu_products = 0
                st.experimental_rerun()
            st.title("Adicionar produto")

            name = st.text_input("Digite o nome do item")

            desc = st.selectbox("Selecione a descrição do item", [
                                'UNIDADE', 'LITRO', 'FRASCO', 'AMPOLA', 'COMPRIMIDO', 'XAROPE', 'CAIXA', 'GOTAS', 'PACOTE', 'QUILO', 'PAR'])
            val = st.number_input("Digite o codigo do item", step=1)

            Add = st.button("Adicionar")
            if Add:
                c.execute("INSERT INTO TABELA_ALMOXARIFADO_PRODUTOS( MODELO_NOME,DESCRICAO,COD_ARQ_LIFE) VALUES ('" +
                          name+"','"+desc+"','" + str(val)+"');")
                connection.commit()

                st.write(":green[Adicionado com sucesso!]")

        if st.session_state.new_form_menu_products == 0:
            st.title("Remover ou adicionar produtos")
            remove = st.button("Remover produto")
            add = st.button("Adicionar produto")
            addDesc = st.button("Adicionar  descrição")
            if remove:
                st.session_state.new_form_menu_products = 2
                st.experimental_rerun()
            if add:
                st.session_state.new_form_menu_products = 3
                st.experimental_rerun()
            if addDesc:
                st.session_state.new_form_menu_products = 4
                st.experimental_rerun()


elif st.session_state.Login2 == 3:
    st.error('Senha ou Usuario esta incorreto')
elif st.session_state.Login2 == 0:
    st.warning('Insira respectivamente o usuario e a senha como solicitado')
